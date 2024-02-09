import openpyxl

def Loader_Code_Emploi(nom_fichier):
    # Charger le fichier Excel
    classeur = openpyxl.load_workbook(nom_fichier)

    # Sélectionner la première feuille
    feuille = classeur.active

    

    # Itérer sur chaque ligne de la feuille
    for ligne in feuille.iter_rows(min_row=2, max_col=feuille.max_column, max_row=feuille.max_row, values_only=True):
        # Traiter chaque cellule de la ligne
        print(ligne, end=" ")  # Vous pouvez remplacer cela par le traitement que vous souhaitez faire pour chaque cellule
        print()  # Passer à la ligne suivante après avoir traité une ligne

    # Fermer le fichier Excel
    classeur.close()

# Exemple d'utilisation
nom_fichier_excel = "Codes emploi par territoire.xlsx"
Loader_Code_Emploi(nom_fichier_excel)